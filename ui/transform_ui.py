import importlib
import os
import sys
import base64
import time
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from IPython.display import display, HTML, clear_output
import ipywidgets as widgets


def detect_columns(file_path, header_row=1):
    """Read Excel header row -> {column_name: column_letter} mapping."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = None
    for s in wb.worksheets:
        if s.sheet_state == 'visible':
            ws = s
            break
    if ws is None:
        ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell.value is not None:
                col_letter = get_column_letter(cell.column)
                mapping[str(cell.value).strip()] = col_letter
    wb.close()
    return mapping


def build_transform_ui(ctx):
    config_state = ctx['config_state']
    file_state = ctx['file_state']
    validator = ctx['validator']
    OUTPUT_DIR = ctx['OUTPUT_DIR']
    UPLOADS_DIR = ctx['UPLOADS_DIR']

    out_transform = widgets.Output()
    out_transform_status = widgets.Output()

    # Live progress log widget
    progress_log = widgets.HTML(value='')

    def log_progress(msg):
        ts = time.strftime('%H:%M:%S')
        current = progress_log.value
        progress_log.value = current + f"<div style='font-family:monospace; font-size:12px; color:#555;'>[{ts}] {msg}</div>"

    # Gate check
    pass_count, fail_count = validator.get_stats()
    gate_ok = (fail_count == 0 and pass_count > 0)

    if not gate_ok:
        with out_transform_status:
            display(HTML(
                "<div style='padding:12px; background:#ffeeba; border-left:4px solid #e74c3c; margin:10px 0;'>"
                f"<b style='color:#e74c3c'>Transformation blocked.</b> "
                f"Validation has {fail_count} failed row(s) and {pass_count} passed row(s).<br>"
                "Fix all validation errors first, then re-run this cell."
                "</div>"
            ))

    # Auto-populated info (read-only display)
    info_html = "<h3>Transformation Configuration</h3>"
    if config_state.get('configured'):
        levels = config_state.get('level_columns', [])
        targets = config_state.get('target_columns', [])
        fac = config_state.get('facility_col', '')
        mapping = config_state.get('alignment_mapping', {})
        b_file = file_state.get('boundary_file', 'N/A')
        f_file = file_state.get('facility_file', 'N/A')
        info_html += f"""
        <div style='padding:10px; background:#e8f4fd; border-radius:5px; margin-bottom:10px;'>
            <b>From Validation Config (auto-populated):</b><br>
            Boundary levels: {', '.join(levels)}<br>
            Target columns: {', '.join(targets) or 'none'}<br>
            Facility column: {fac or 'none'}<br>
            Mapping: {', '.join(f'{k}&rarr;{v}' for k,v in mapping.items()) or 'none'}<br>
            Boundary file: {os.path.basename(b_file) if b_file else 'N/A'}<br>
            Facility file: {os.path.basename(f_file) if f_file else 'N/A'}<br>
            <em>Column letters and province name will be auto-detected from uploaded files.</em>
        </div>"""

    # Settings widgets
    style = {'description_width': '140px'}
    w_db_name = widgets.Text(
        value='microplan.db', description='DB Filename:',
        style=style, layout=widgets.Layout(width='400px'))
    w_country_code = widgets.Text(
        value='mz', description='Country Code:',
        style=style, layout=widgets.Layout(width='400px'))
    w_project_name = widgets.Text(
        value='IRS', description='Project Name:',
        style=style, layout=widgets.Layout(width='400px'))

    # Campaign date pickers
    today = datetime.date.today()
    w_campaign_start = widgets.DatePicker(
        description='Campaign Start:',
        value=today,
        style=style, layout=widgets.Layout(width='300px'))
    w_campaign_end = widgets.DatePicker(
        description='Campaign End:',
        value=today + datetime.timedelta(days=30),
        style=style, layout=widgets.Layout(width='300px'))

    w_boundary_start_row = widgets.IntText(
        value=1, description='Boundary Header Row:',
        style=style, layout=widgets.Layout(width='250px'))
    w_facility_start_row = widgets.IntText(
        value=1, description='Facility Header Row:',
        style=style, layout=widgets.Layout(width='250px'))

    # Checklist upload (optional)
    w_checklist = widgets.FileUpload(
        accept='.xlsx,.xls', multiple=False, description='Checklist File')

    btn_transform = widgets.Button(
        description='TRANSFORM', button_style='success',
        layout=widgets.Layout(width='150px', height='40px'),
        disabled=not gate_ok)

    # Store w_db_name in ctx so export_ui can read it
    ctx['widgets']['w_db_name'] = w_db_name

    def on_transform(b):
        progress_log.value = ''
        btn_transform.disabled = True
        btn_transform.description = 'RUNNING...'

        with out_transform:
            clear_output(wait=True)

            if not config_state.get('configured'):
                display(HTML("<p style='color:red'><b>Column configuration not saved!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return
            if not file_state.get('boundary_file') or not file_state.get('facility_file'):
                display(HTML("<p style='color:red'><b>Boundary and Facility files required!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return

            # Validate dates
            if not w_campaign_start.value or not w_campaign_end.value:
                display(HTML("<p style='color:red'><b>Campaign start and end dates are required!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return
            if w_campaign_end.value < w_campaign_start.value:
                display(HTML("<p style='color:red'><b>Campaign end date must be after start date!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return

            try:
                log_progress("Starting transformation...")
                level_columns = config_state.get('level_columns', [])
                target_columns = config_state.get('target_columns', [])
                header_row = w_boundary_start_row.value

                # Auto-detect column letters from boundary file header
                log_progress("Detecting column letters from Excel headers...")
                header_map = detect_columns(file_state['boundary_file'], header_row)

                boundary_columns = {}
                missing = []
                for i, level_name in enumerate(level_columns):
                    if i == 0:
                        continue
                    level_num = i + 1
                    if level_name in header_map:
                        boundary_columns[level_num] = header_map[level_name]
                    else:
                        missing.append(level_name)

                target_column_letters = {}
                for t_name in target_columns:
                    if t_name in header_map:
                        target_column_letters[t_name] = header_map[t_name]
                    else:
                        missing.append(t_name)

                if missing:
                    display(HTML(
                        f"<p style='color:red'><b>Columns not found in boundary file header row {header_row}:</b> "
                        f"{', '.join(missing)}<br>"
                        f"Found headers: {', '.join(header_map.keys())}</p>"))
                    btn_transform.disabled = False
                    btn_transform.description = 'TRANSFORM'
                    return

                log_progress(f"Detected {len(boundary_columns)} boundary columns, {len(target_column_letters)} target columns.")

                # Auto-detect province name from the first data row
                log_progress("Detecting province name from data...")
                province_col = boundary_columns.get(2)
                province_name = ''
                if province_col:
                    wb_tmp = openpyxl.load_workbook(file_state['boundary_file'], data_only=True, read_only=True)
                    ws_tmp = None
                    for s in wb_tmp.worksheets:
                        if s.sheet_state == 'visible':
                            ws_tmp = s
                            break
                    if ws_tmp is None:
                        ws_tmp = wb_tmp.active
                    for row in ws_tmp.iter_rows(min_row=header_row + 1, max_row=header_row + 1):
                        for cell in row:
                            if get_column_letter(cell.column) == province_col and cell.value:
                                province_name = str(cell.value).strip()
                                break
                    wb_tmp.close()

                if not province_name:
                    display(HTML(
                        "<p style='color:red'><b>Could not detect province name from boundary file.</b></p>"))
                    btn_transform.disabled = False
                    btn_transform.description = 'TRANSFORM'
                    return

                log_progress(f"Province detected: {province_name}")

                # Format campaign dates
                date_fmt = '%d/%m/%Y'
                campaign_start = w_campaign_start.value.strftime(date_fmt)
                campaign_end = w_campaign_end.value.strftime(date_fmt)
                log_progress(f"Campaign: {campaign_start} to {campaign_end}")

                # Show detected mapping
                detect_html = "<div style='padding:8px; background:#f0f8e8; border-radius:4px; margin:8px 0; font-size:12px;'>"
                detect_html += f"<b>Auto-detected:</b> Province = <b>{province_name}</b><br>"
                detect_html += "Columns: " + ", ".join(
                    f"{name}={boundary_columns.get(i+1, '?')}"
                    for i, name in enumerate(level_columns) if i > 0
                )
                if target_column_letters:
                    detect_html += "<br>Targets: " + ", ".join(
                        f"{n}={c}" for n, c in target_column_letters.items()
                    )
                detect_html += f"<br>Campaign: {campaign_start} &rarr; {campaign_end}"
                detect_html += "</div>"
                display(HTML(detect_html))

                # DB path inside output folder
                db_filename = w_db_name.value.strip() or 'microplan.db'
                db_path_in_output = os.path.join(OUTPUT_DIR, db_filename)

                # Build user inputs
                user_inputs = {
                    'db_name': db_path_in_output,
                    'country_code': w_country_code.value.strip() or 'mz',
                    'province_name': province_name,
                    'province_code': '',
                    'project_name': w_project_name.value.strip() or 'IRS',
                    'boundary_start_row': header_row,
                    'facility_start_row': w_facility_start_row.value,
                    'campaign_start_date': campaign_start,
                    'campaign_end_date': campaign_end,
                }

                config_with_cols = dict(config_state)
                config_with_cols['boundary_columns'] = boundary_columns
                config_with_cols['target_column_letters'] = target_column_letters

                # Reload modules (clears stale state)
                log_progress("Reloading modules...")

                if 'models.db.Base' in sys.modules:
                    sys.modules['models.db.Base'].Base.metadata.clear()

                _modules_to_reload = [
                    'constants.constants',
                    'models.db.Base',
                    'models.db.Boundary',
                    'models.db.Facility',
                    'models.db',
                    'models',
                    'utils.common',
                    'utils.boundary',
                    'utils.facility',
                    'utils',
                    'transform',
                ]
                for mod_name in _modules_to_reload:
                    if mod_name in sys.modules:
                        importlib.reload(sys.modules[mod_name])

                # Build and apply config AFTER reload
                log_progress("Applying configuration...")
                from constants.constants import TransformConfig
                cfg = TransformConfig.from_notebook(config_with_cols, user_inputs)
                cfg.apply_to_module()

                # Import and run
                from transform import run_transform

                checklist_path = None
                if w_checklist.value:
                    files = w_checklist.value
                    info = files[0] if isinstance(files, tuple) else list(files.values())[0]
                    name = info.name if hasattr(info, 'name') else info['name']
                    content = info.content if hasattr(info, 'content') else info['content']
                    checklist_path = os.path.join(UPLOADS_DIR, name)
                    with open(checklist_path, 'wb') as f:
                        f.write(content)
                elif os.path.exists('checklist_targets.xlsx'):
                    checklist_path = 'checklist_targets.xlsx'

                log_progress("Starting data transformation...")
                result = run_transform(
                    boundary_file=file_state['boundary_file'],
                    facility_file=file_state['facility_file'],
                    checklist_file=checklist_path,
                    progress=log_progress,
                )

                # Show results + download
                db_path = result['db_path']
                html = f"""
                <div style='padding:12px; background:#d4edda; border-left:4px solid #27ae60; margin:10px 0;'>
                    <b style='color:#27ae60; font-size:16px;'>Transformation Complete!</b><br><br>
                    <table style='font-size:13px;'>
                        <tr><td><b>Database:</b></td><td>{db_path}</td></tr>
                        <tr><td><b>Province:</b></td><td>{province_name}</td></tr>
                        <tr><td><b>Campaign:</b></td><td>{campaign_start} &rarr; {campaign_end}</td></tr>
                        <tr><td><b>Boundaries created:</b></td><td>{result['boundaries_count']}</td></tr>
                        <tr><td><b>Facilities created:</b></td><td>{result['facilities_count']}</td></tr>
                    </table>
                </div>"""

                if os.path.exists(db_path):
                    with open(db_path, 'rb') as f:
                        b64 = base64.b64encode(f.read()).decode()
                    html += (
                        f'<a href="data:application/octet-stream;base64,{b64}" '
                        f'download="{os.path.basename(db_path)}" '
                        f'style="display:inline-block; padding:10px 20px; background:#2196F3; color:white; '
                        f'text-decoration:none; border-radius:4px; margin:8px 4px; font-weight:bold;">'
                        f'Download {os.path.basename(db_path)}</a>'
                    )

                display(HTML(html))

            except Exception as ex:
                import traceback
                log_progress(f"ERROR: {ex}")
                display(HTML(
                    f"<div style='padding:10px; background:#ffc7ce; border-radius:5px;'>"
                    f"<b style='color:red'>Transformation Error:</b><br>"
                    f"<pre>{traceback.format_exc()}</pre></div>"
                ))
            finally:
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'

    btn_transform.on_click(on_transform)

    return widgets.VBox([
        widgets.HTML(info_html),
        out_transform_status,
        widgets.HTML("<b>Settings:</b>"),
        w_db_name, w_country_code, w_project_name,
        widgets.HTML("<b>Campaign Dates:</b>"),
        w_campaign_start, w_campaign_end,
        widgets.HTML("<b>Excel Row Settings:</b>"),
        w_boundary_start_row, w_facility_start_row,
        widgets.HTML("<b>Optional Checklist File:</b>"), w_checklist,
        widgets.HTML("<br>"),
        btn_transform,
        widgets.HTML("<b>Progress:</b>"),
        progress_log,
        out_transform,
    ])
