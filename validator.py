import os
import re
import pandas as pd
import yaml
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment


class Validator:
    """Validates microplan Excel/CSV files for data quality issues."""

    def __init__(self, config_path='validation_config.yaml'):
        self.config = self._load_config(config_path)

        self.boundary_columns = []
        self.facility_columns = []
        self.target_columns = []
        self.user_columns = []
        self.num_targets = 0

        self.special_allowed = self._get_config_list(['validation_rules', 'special_characters', 'allowed_special_chars'])

        self.rules_enabled = {
            'non_zero_targets': self._get_config_bool(['validation_rules', 'non_zero_targets', 'enabled']),
            'naming_convention': self._get_config_bool(['validation_rules', 'naming_convention', 'enabled']),
            'boundary_alignment': self._get_config_bool(['validation_rules', 'boundary_alignment', 'enabled']),
            'unique_names': self._get_config_bool(['validation_rules', 'unique_names', 'enabled']),
            'user_mapping': self._get_config_bool(['validation_rules', 'user_mapping', 'enabled']),
            'no_missing_entries': self._get_config_bool(['validation_rules', 'no_missing_entries', 'enabled']),
            'special_characters': self._get_config_bool(['validation_rules', 'special_characters', 'enabled']),
            'hierarchy_check': self._get_config_bool(['validation_rules', 'hierarchy_check', 'enabled'])
        }

        self.rules_severity = {
            'non_zero_targets': self._get_config_value(['validation_rules', 'non_zero_targets', 'severity'], 'error'),
            'naming_convention': self._get_config_value(['validation_rules', 'naming_convention', 'severity'], 'warning'),
            'boundary_alignment': self._get_config_value(['validation_rules', 'boundary_alignment', 'severity'], 'error'),
            'unique_names': self._get_config_value(['validation_rules', 'unique_names', 'severity'], 'error'),
            'user_mapping': self._get_config_value(['validation_rules', 'user_mapping', 'severity'], 'warning'),
            'no_missing_entries': self._get_config_value(['validation_rules', 'no_missing_entries', 'severity'], 'error'),
            'special_characters': self._get_config_value(['validation_rules', 'special_characters', 'severity'], 'error'),
            'hierarchy_check': self._get_config_value(['validation_rules', 'hierarchy_check', 'severity'], 'error')
        }

        self.hierarchy_auto_detect_root = self._get_config_bool(['validation_rules', 'hierarchy_check', 'auto_detect_root'])
        self.hierarchy_root_threshold_rows = self._get_config_value(['validation_rules', 'hierarchy_check', 'root_threshold_rows'], 5)
        self.hierarchy_root_threshold_percent = self._get_config_value(['validation_rules', 'hierarchy_check', 'root_threshold_percent'], 0.1)

        self.row_status = {}
        self.file_data = {}
        self.output_files = []
        self.alignment_mapping = {}

    def _load_config(self, config_path):
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f)
        except Exception as e:
            print(f"Warning: Could not load config '{config_path}': {e}")
        return {}

    def _get_config_list(self, keys):
        try:
            value = self.config
            for key in keys:
                value = value[key]
            return value if isinstance(value, list) else []
        except (KeyError, TypeError):
            return []

    def _get_config_bool(self, keys):
        try:
            value = self.config
            for key in keys:
                value = value[key]
            return bool(value)
        except (KeyError, TypeError):
            return False

    def _get_config_value(self, keys, default=None):
        try:
            value = self.config
            for key in keys:
                value = value[key]
            return value
        except (KeyError, TypeError):
            return default

    def set_columns(self, boundary_cols=None, facility_cols=None, target_cols=None, user_cols=None, num_targets=0):
        self.boundary_columns = boundary_cols or []
        self.facility_columns = facility_cols or []
        self.target_columns = target_cols or []
        self.user_columns = user_cols or []
        self.num_targets = num_targets

    def set_alignment_mapping(self, mapping):
        self.alignment_mapping = mapping or {}

    def find_cols(self, df, col_list):
        """Find columns from dataframe that match the given list (case-insensitive)."""
        df_cols_lower = {str(c).lower(): c for c in df.columns}
        found = []
        for col in col_list:
            col_lower = str(col).lower().strip()
            if col_lower in df_cols_lower:
                found.append(df_cols_lower[col_lower])
        return found

    def find_parent_col(self, df):
        for c in df.columns:
            if 'parent' in str(c).lower():
                return c
        return None

    def is_csv(self, filepath):
        return filepath.lower().endswith('.csv')

    def init_row_status(self, df, sheet):
        self.row_status[sheet] = {}
        for idx in df.index:
            self.row_status[sheet][idx] = {'status': 'PASS', 'errors': []}

    def mark_row_error(self, sheet, row, error_msg):
        if sheet in self.row_status and row in self.row_status[sheet]:
            self.row_status[sheet][row]['status'] = 'FAIL'
            self.row_status[sheet][row]['errors'].append(error_msg)

    def check_non_zero(self, df, sheet):
        """Validate target values are non-zero and rounded."""
        if not self.rules_enabled['non_zero_targets']:
            return []

        issues = []
        target_cols = self.find_cols(df, self.target_columns)

        for col in target_cols:
            for idx, val in df[col].items():
                if pd.notna(val):
                    try:
                        n = float(val)
                        if n == 0:
                            sev = self.rules_severity['non_zero_targets']
                            issues.append({
                                'rule': 'Non-Zero Targets', 'severity': sev, 'sheet': sheet,
                                'column': col, 'row': idx + 2, 'value': val, 'message': 'Zero value'
                            })
                            if sev == 'error': self.mark_row_error(sheet, idx, f'Zero value in {col}')
                        elif n != round(n):
                            issues.append({
                                'rule': 'Non-Zero Targets', 'severity': 'warning', 'sheet': sheet,
                                'column': col, 'row': idx + 2, 'value': val,
                                'message': f'Not rounded (should be {round(n)})'
                            })
                    except:
                        pass
        return issues

    def check_naming(self, df, sheet):
        """Check for consistent naming convention across name columns."""
        if not self.rules_enabled['naming_convention']:
            return []

        issues = []
        name_cols = self.find_cols(df, self.boundary_columns) + self.find_cols(df, self.facility_columns)
        names = [(c, i, str(v).strip()) for c in name_cols for i, v in df[c].items() if pd.notna(v)]

        if names:
            cases = {'upper': 0, 'lower': 0, 'title': 0, 'mixed': 0}
            for _, _, n in names:
                if n.isupper(): cases['upper'] += 1
                elif n.islower(): cases['lower'] += 1
                elif n.istitle(): cases['title'] += 1
                else: cases['mixed'] += 1

            dominant = max(cases, key=cases.get)
            cnt = 0
            sev = self.rules_severity['naming_convention']
            for c, i, n in names:
                if not (n.isupper() or n.islower() or n.istitle()) and cnt < 10:
                    issues.append({
                        'rule': 'Naming Convention', 'severity': sev, 'sheet': sheet,
                        'column': c, 'row': i + 2, 'value': n[:40],
                        'message': f'Inconsistent case (dominant: {dominant})'
                    })
                    cnt += 1
        return issues

    def check_alignment(self, b_df, f_df, b_sheet, f_sheet):
        """Check that facility data exists in boundary file."""
        if not self.rules_enabled['boundary_alignment'] or not self.alignment_mapping:
            return []

        issues = []
        b_cols_lower = {str(c).lower(): c for c in b_df.columns}
        f_cols_lower = {str(c).lower(): c for c in f_df.columns}

        valid_values = {}
        col_mapping = {}

        for f_col, b_col in self.alignment_mapping.items():
            actual_f_col = f_cols_lower.get(str(f_col).lower())
            actual_b_col = b_cols_lower.get(str(b_col).lower())

            if actual_f_col and actual_b_col:
                col_mapping[actual_f_col] = actual_b_col
                valid_values[actual_b_col] = set(
                    b_df[actual_b_col].dropna().astype(str).str.strip().str.lower().unique()
                )

        sev = self.rules_severity['boundary_alignment']
        for idx in f_df.index:
            for f_col, b_col in col_mapping.items():
                val = f_df.loc[idx, f_col]
                if pd.notna(val):
                    val_str = str(val).strip()
                    if val_str and val_str.lower() not in valid_values[b_col]:
                        issues.append({
                            'rule': 'Boundary Alignment', 'severity': sev, 'sheet': f_sheet,
                            'column': f_col, 'row': idx + 2, 'value': val_str[:40],
                            'message': f'"{val_str}" not found in {b_col}'
                        })
                        if sev == 'error': self.mark_row_error(f_sheet, idx, f'{f_col} "{val_str}" not in boundary')
        return issues

    def check_unique(self, df, sheet):
        """Check for duplicate names under same parent hierarchy."""
        if not self.rules_enabled['unique_names']:
            return []

        issues = []
        sev = self.rules_severity['unique_names']
        boundary_cols = self.find_cols(df, self.boundary_columns)
        facility_cols = self.find_cols(df, self.facility_columns)

        # Detect if this is a boundary file (has 4+ hierarchy columns) vs facility file
        # Facility files typically have few columns like [Facility Name, District, State]
        is_boundary_file = len(boundary_cols) >= 4

        # Only check boundary hierarchy duplicates for actual boundary files
        if boundary_cols and is_boundary_file:
            last_col = boundary_cols[-1]
            parent_cols = boundary_cols[:-1]

            if parent_cols:
                try:
                    df['_parent_key'] = df[parent_cols].astype(str).agg('|'.join, axis=1)
                    for parent_key, group in df.groupby('_parent_key')[last_col]:
                        vals = group.dropna().astype(str).str.strip()
                        seen = {}
                        for idx, val in vals.items():
                            if val in seen:
                                parent_display = ' > '.join([str(df.loc[idx, c]) for c in parent_cols])
                                issues.append({
                                    'rule': 'Unique Names', 'severity': sev, 'sheet': sheet,
                                    'column': last_col, 'row': idx + 2, 'value': val[:40],
                                    'message': f'Duplicate under "{parent_display}" (also row {seen[val] + 2})'
                                })
                                if sev == 'error': self.mark_row_error(sheet, idx, f'Duplicate {last_col} "{val}"')
                            else:
                                seen[val] = idx
                    df.drop('_parent_key', axis=1, inplace=True, errors='ignore')
                except:
                    pass
            else:
                vals = df[last_col].dropna().astype(str).str.strip()
                seen = {}
                for idx, val in vals.items():
                    if val in seen:
                        issues.append({
                            'rule': 'Unique Names', 'severity': sev, 'sheet': sheet,
                            'column': last_col, 'row': idx + 2, 'value': val[:40],
                            'message': f'Duplicate (also row {seen[val] + 2})'
                        })
                        if sev == 'error': self.mark_row_error(sheet, idx, f'Duplicate {last_col}: {val}')
                    else:
                        seen[val] = idx

        # Check facility columns for duplicates
        for fac_col in facility_cols:
            # For boundary files, group by last boundary column
            # For facility files, check global duplicates
            if is_boundary_file and boundary_cols:
                group_col = boundary_cols[-1]
                try:
                    for parent_val, group in df.groupby(group_col)[fac_col]:
                        vals = group.dropna().astype(str).str.strip()
                        seen = {}
                        for idx, val in vals.items():
                            if val in seen:
                                issues.append({
                                    'rule': 'Unique Names', 'severity': sev, 'sheet': sheet,
                                    'column': fac_col, 'row': idx + 2, 'value': val[:40],
                                    'message': f'Duplicate facility under "{parent_val}" (also row {seen[val] + 2})'
                                })
                                if sev == 'error': self.mark_row_error(sheet, idx, f'Duplicate facility "{val}"')
                            else:
                                seen[val] = idx
                except:
                    pass
            else:
                # Facility file - check for global duplicates in facility column
                vals = df[fac_col].dropna().astype(str).str.strip()
                seen = {}
                for idx, val in vals.items():
                    if val in seen:
                        issues.append({
                            'rule': 'Unique Names', 'severity': sev, 'sheet': sheet,
                            'column': fac_col, 'row': idx + 2, 'value': val[:40],
                            'message': f'Duplicate facility (also row {seen[val] + 2})'
                        })
                        if sev == 'error': self.mark_row_error(sheet, idx, f'Duplicate facility "{val}"')
                    else:
                        seen[val] = idx
        return issues

    def check_users(self, df, sheet):
        """Check for duplicate phone/contact numbers."""
        if not self.rules_enabled['user_mapping']:
            return []

        issues = []
        sev = self.rules_severity['user_mapping']
        phone_cols = [c for c in self.find_cols(df, self.user_columns)
                      if any(p in str(c).lower() for p in ['mobile', 'phone', 'contact'])]

        for col in phone_cols:
            if col in df.columns:
                vals = df[col].dropna().astype(str).str.replace(r'[\s\-\(\)]', '', regex=True)
                seen = {}
                for idx, val in vals.items():
                    if val:
                        if val in seen:
                            issues.append({
                                'rule': 'User Mapping', 'severity': sev, 'sheet': sheet,
                                'column': col, 'row': idx + 2, 'value': val,
                                'message': f'Duplicate contact (also row {seen[val] + 2})'
                            })
                        else:
                            seen[val] = idx
        return issues

    def check_missing(self, df, sheet):
        """Check for blank values in required fields."""
        if not self.rules_enabled['no_missing_entries']:
            return []

        issues = []
        sev = self.rules_severity['no_missing_entries']
        check_cols = set()
        check_cols.update(self.find_cols(df, self.boundary_columns))
        check_cols.update(self.find_cols(df, self.facility_columns))
        check_cols.update(self.find_cols(df, self.target_columns))

        for col in check_cols:
            nulls = df[col].isna() | (df[col].astype(str).str.strip() == '')
            for idx in df[nulls].index:
                issues.append({
                    'rule': 'No Missing Entries', 'severity': sev, 'sheet': sheet,
                    'column': col, 'row': idx + 2, 'value': 'BLANK', 'message': 'Missing value'
                })
                if sev == 'error': self.mark_row_error(sheet, idx, f'Missing value in {col}')
        return issues

    def check_special(self, df, sheet):
        """Check for unusual special characters in names."""
        if not self.rules_enabled['special_characters']:
            return []

        issues = []
        sev = self.rules_severity['special_characters']
        pattern = f'[^a-zA-Z0-9{re.escape("".join(self.special_allowed))}]'
        cnt = 0

        name_cols = set()
        name_cols.update(self.find_cols(df, self.boundary_columns))
        name_cols.update(self.find_cols(df, self.facility_columns))

        for col in name_cols:
            for idx, val in df[col].items():
                if pd.notna(val) and cnt < 20:
                    s = str(val).strip()
                    chars = re.findall(pattern, s)
                    if chars:
                        issues.append({
                            'rule': 'Special Characters', 'severity': sev, 'sheet': sheet,
                            'column': col, 'row': idx + 2, 'value': s[:40],
                            'message': f'Found: {list(set(chars))}'
                        })
                        if sev == 'error': self.mark_row_error(sheet, idx, f'Special chars in {col}')
                        cnt += 1
        return issues

    def check_hierarchy(self, df, sheet):
        """Check that parent references exist in the data."""
        if not self.rules_enabled['hierarchy_check']:
            return []

        issues = []
        parent_col = self.find_parent_col(df)

        if not parent_col or parent_col not in df.columns:
            return issues

        boundary_cols = self.find_cols(df, self.boundary_columns)

        first_col = df.columns[0] if len(df.columns) > 0 else None
        code_col = None
        if first_col and any(p in str(first_col).lower() for p in ['code', 'id', 'boundary', 'key']):
            code_col = first_col

        valid_parents = set()
        if code_col:
            valid_parents.update(df[code_col].dropna().astype(str).str.strip().unique())
        for col in boundary_cols:
            valid_parents.update(df[col].dropna().astype(str).str.strip().unique())

        detected_roots = set()
        if self.hierarchy_auto_detect_root:
            parent_counts = df[parent_col].dropna().astype(str).str.strip().value_counts()
            for parent_val, count in parent_counts.items():
                if parent_val not in valid_parents:
                    if count > self.hierarchy_root_threshold_rows or count > len(df) * self.hierarchy_root_threshold_percent:
                        detected_roots.add(parent_val)

        sev = self.rules_severity['hierarchy_check']
        for idx, parent_val in df[parent_col].items():
            if pd.notna(parent_val):
                parent_str = str(parent_val).strip()
                if parent_str and parent_str not in valid_parents and parent_str not in detected_roots:
                    issues.append({
                        'rule': 'Hierarchy Check', 'severity': sev, 'sheet': sheet,
                        'column': parent_col, 'row': idx + 2, 'value': parent_str[:40],
                        'message': f'Parent "{parent_str}" not found'
                    })
                    if sev == 'error': self.mark_row_error(sheet, idx, f'Invalid parent: {parent_str}')
        return issues

    def check_columns_exist(self, df, sheet):
        """Check if configured columns exist in the dataframe."""
        issues = []
        df_cols_lower = {str(c).lower() for c in df.columns}

        # Determine which columns are found
        boundary_found = [c for c in self.boundary_columns if c and str(c).lower().strip() in df_cols_lower]
        facility_found = [c for c in self.facility_columns if c and str(c).lower().strip() in df_cols_lower]
        target_found = [c for c in self.target_columns if c and str(c).lower().strip() in df_cols_lower]

        # Detect file type based on which columns are present
        # If multiple boundary hierarchy columns found -> likely boundary file
        # If facility columns found but few boundary columns -> likely facility file
        is_boundary_file = len(boundary_found) >= 3
        is_facility_file = len(facility_found) > 0 and len(boundary_found) < 3

        # Build list of columns to check based on file type
        cols_to_check = []

        if is_boundary_file:
            # For boundary files: check boundary columns and target columns
            cols_to_check.extend(self.boundary_columns)
            cols_to_check.extend(self.target_columns)
        elif is_facility_file:
            # For facility files: check facility columns only
            cols_to_check.extend(self.facility_columns)
        else:
            # Unknown file type: check all columns that might be relevant
            # Only report missing if NONE from a category are found
            if not boundary_found and self.boundary_columns:
                cols_to_check.extend(self.boundary_columns)
            if not facility_found and self.facility_columns:
                cols_to_check.extend(self.facility_columns)
            if not target_found and self.target_columns:
                cols_to_check.extend(self.target_columns)

        for col in cols_to_check:
            if col and str(col).lower().strip() not in df_cols_lower:
                issues.append({
                    'rule': 'Column Not Found', 'severity': 'error', 'sheet': sheet,
                    'column': col, 'row': '-', 'value': '-',
                    'message': f'Column "{col}" not found. Available: {list(df.columns)[:5]}...'
                })
        return issues

    def validate_df(self, df, sheet):
        """Run all validation checks on a dataframe."""
        self.init_row_status(df, sheet)
        issues = []
        issues.extend(self.check_columns_exist(df, sheet))
        issues.extend(self.check_non_zero(df, sheet))
        issues.extend(self.check_naming(df, sheet))
        issues.extend(self.check_unique(df, sheet))
        issues.extend(self.check_users(df, sheet))
        issues.extend(self.check_missing(df, sheet))
        issues.extend(self.check_special(df, sheet))
        issues.extend(self.check_hierarchy(df, sheet))
        return issues

    def read_file(self, filepath):
        """Read Excel or CSV file into dict of dataframes."""
        if self.is_csv(filepath):
            return {os.path.basename(filepath): pd.read_csv(filepath)}
        else:
            xls = pd.ExcelFile(filepath)
            return {sheet: pd.read_excel(xls, sheet_name=sheet)
                    for sheet in xls.sheet_names
                    if not pd.read_excel(xls, sheet_name=sheet).empty}

    def validate_file(self, filepath, b_sheet=None, f_sheet=None):
        """Validate a file and return issues with summary."""
        issues = []
        b_df, f_df, b_name, f_name = None, None, None, None

        if filepath not in self.file_data:
            self.file_data[filepath] = {}

        try:
            fname = os.path.basename(filepath)
            sheets_data = self.read_file(filepath)

            for sheet, df in sheets_data.items():
                if df.empty:
                    continue

                label = f"{fname} - {sheet}" if not self.is_csv(filepath) else fname
                issues.extend(self.validate_df(df, label))

                df_with_status = df.copy()
                df_with_status['VALIDATION_STATUS'] = 'PASS'
                df_with_status['VALIDATION_ERRORS'] = ''

                if label in self.row_status:
                    for idx, status_info in self.row_status[label].items():
                        if idx in df_with_status.index:
                            df_with_status.loc[idx, 'VALIDATION_STATUS'] = status_info['status']
                            df_with_status.loc[idx, 'VALIDATION_ERRORS'] = '; '.join(status_info['errors'])

                self.file_data[filepath][sheet] = df_with_status

                sl = sheet.lower()
                if b_sheet and b_sheet.lower() == sl:
                    b_df, b_name = df, label
                elif f_sheet and f_sheet.lower() == sl:
                    f_df, f_name = df, label

            if b_df is not None and f_df is not None:
                issues.extend(self.check_alignment(b_df, f_df, b_name, f_name))

        except Exception as e:
            issues.append({
                'rule': 'File Error', 'severity': 'error', 'sheet': filepath,
                'column': '-', 'row': '-', 'value': '-', 'message': str(e)
            })

        return issues, self.summarize(issues)

    def summarize(self, issues):
        summary = {
            'total': len(issues),
            'errors': len([i for i in issues if i['severity'] == 'error']),
            'warnings': len([i for i in issues if i['severity'] == 'warning']),
            'by_rule': defaultdict(int)
        }
        for i in issues:
            summary['by_rule'][i['rule']] += 1
        return summary

    def save_validated_files(self, output_folder='error'):
        """Save validated files with color-coded status."""
        self.output_files = []
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for filepath, sheets_data in self.file_data.items():
            if not sheets_data:
                continue

            name_part = os.path.splitext(os.path.basename(filepath))[0]
            output_path = os.path.join(output_folder, f"{name_part}_VALIDATED_{ts}.xlsx")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_data.items():
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

                for sheet_name in writer.book.sheetnames:
                    self._apply_colors(writer.book[sheet_name], header_fill, header_font, green_fill, red_fill)

            self.output_files.append(output_path)
        return self.output_files

    def _apply_colors(self, ws, header_fill, header_font, green_fill, red_fill):
        status_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == 'VALIDATION_STATUS':
                status_col = col_idx
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        if not status_col:
            return

        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=status_col)
            if status_cell.value == 'PASS':
                status_cell.fill = green_fill
                status_cell.font = Font(bold=True, color='006100')
            elif status_cell.value == 'FAIL':
                status_cell.fill = red_fill
                status_cell.font = Font(bold=True, color='9C0006')

    def get_stats(self):
        pass_count = sum(1 for sd in self.row_status.values() for rd in sd.values() if rd['status'] == 'PASS')
        fail_count = sum(1 for sd in self.row_status.values() for rd in sd.values() if rd['status'] == 'FAIL')
        return pass_count, fail_count

    def reset(self):
        self.row_status = {}
        self.file_data = {}
        self.output_files = []
        self.alignment_mapping = {}
