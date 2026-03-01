import importlib
import os
import sys
import base64
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from IPython.display import display, HTML, clear_output
import ipywidgets as widgets
from utils.common import detect_columns


def build_transform_ui(ctx):
    config_state = ctx['config_state']
    file_state = ctx['file_state']
    validator = ctx['validator']
    OUTPUT_DIR = ctx['OUTPUT_DIR']
    UPLOADS_DIR = ctx['UPLOADS_DIR']

    out_transform = widgets.Output()
    out_transform_status = widgets.Output()

    # Progress bar + label
    progress_bar = widgets.IntProgress(
        value=0, min=0, max=1,
        description='',
        bar_style='info',
        style={'description_width': '0px'},
        layout=widgets.Layout(width='500px', visibility='hidden'),
    )
    progress_label = widgets.HTML(value='')
    progress_box = widgets.HBox([progress_bar, progress_label])

    def update_progress(current, total, msg):
        progress_bar.max = max(total, 1)
        progress_bar.value = current
        progress_bar.layout.visibility = 'visible'
        pct = int(current / max(total, 1) * 100)
        progress_label.value = (
            f"<span style='margin-left:10px; font-size:13px;'>"
            f"{msg} ({pct}%)</span>"
        )

    # Gate check
    pass_count, fail_count = validator.get_stats()
    gate_ok = (fail_count == 0 and pass_count > 0)

    if not gate_ok:
        with out_transform_status:
            display(HTML(
                "<div style='padding:12px; background:#ffeeba; border-left:4px solid #e74c3c; margin:10px 0;'>"
                f"<b style='color:#e74c3c'>Transformation blocked.</b> "
                f"Validation has {fail_count} failed row(s) and {pass_count} passed row(s).<br>"
                "Fix all validation errors first, then re-run this cell."
                "</div>"
            ))

    info_html = "<h3>Transformation Configuration</h3>"

    # Settings widgets
    style = {'description_width': '140px'}
    w_db_name = widgets.Text(
        value='microplan.db', description='DB Filename:',
        style=style, layout=widgets.Layout(width='400px'))
    w_country_code = widgets.Text(
        value='mz', description='Country Code:',
        style=style, layout=widgets.Layout(width='400px'))
    w_project_name = widgets.Text(
        value='IRS', description='Project Name:',
        style=style, layout=widgets.Layout(width='400px'))

    # Campaign date pickers
    today = datetime.date.today()
    w_campaign_start = widgets.DatePicker(
        description='Campaign Start:',
        value=today,
        style=style, layout=widgets.Layout(width='300px'))
    w_campaign_end = widgets.DatePicker(
        description='Campaign End:',
        value=today + datetime.timedelta(days=30),
        style=style, layout=widgets.Layout(width='300px'))

    w_boundary_start_row = widgets.IntText(
        value=1, description='Boundary Header Row:',
        style=style, layout=widgets.Layout(width='250px'))
    w_facility_start_row = widgets.IntText(
        value=1, description='Facility Header Row:',
        style=style, layout=widgets.Layout(width='250px'))

    btn_transform = widgets.Button(
        description='TRANSFORM', button_style='success',
        layout=widgets.Layout(width='150px', height='40px'),
        disabled=not gate_ok)

    # Store w_db_name in ctx so export_ui can read it
    ctx['widgets']['w_db_name'] = w_db_name

    def on_transform(b):
        progress_bar.value = 0
        progress_bar.bar_style = 'info'
        progress_bar.layout.visibility = 'hidden'
        progress_label.value = ''
        btn_transform.disabled = True
        btn_transform.description = 'RUNNING...'

        with out_transform:
            clear_output(wait=True)

            if not config_state.get('configured'):
                display(HTML("<p style='color:red'><b>Column configuration not saved!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return
            if not file_state.get('boundary_file') or not file_state.get('facility_file'):
                display(HTML("<p style='color:red'><b>Boundary and Facility files required!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return

            # Validate dates
            if not w_campaign_start.value or not w_campaign_end.value:
                display(HTML("<p style='color:red'><b>Campaign start and end dates are required!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return
            if w_campaign_end.value < w_campaign_start.value:
                display(HTML("<p style='color:red'><b>Campaign end date must be after start date!</b></p>"))
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'
                return

            try:
                update_progress(0, 100, "Starting transformation...")
                level_columns = config_state.get('level_columns', [])
                target_columns = config_state.get('target_columns', [])
                header_row = w_boundary_start_row.value

                # Auto-detect column letters from boundary file header
                update_progress(10, 100, "Detecting columns...")
                header_map = detect_columns(file_state['boundary_file'], header_row)

                boundary_columns = {}
                missing = []
                for i, level_name in enumerate(level_columns):
                    if i == 0:
                        continue
                    level_num = i + 1
                    if level_name in header_map:
                        boundary_columns[level_num] = header_map[level_name]
                    else:
                        missing.append(level_name)

                target_column_letters = {}
                for t_name in target_columns:
                    if t_name in header_map:
                        target_column_letters[t_name] = header_map[t_name]
                    else:
                        missing.append(t_name)

                if missing:
                    display(HTML(
                        f"<p style='color:red'><b>Columns not found in boundary file header row {header_row}:</b> "
                        f"{', '.join(missing)}<br>"
                        f"Found headers: {', '.join(header_map.keys())}</p>"))
                    btn_transform.disabled = False
                    btn_transform.description = 'TRANSFORM'
                    return

                update_progress(20, 100, f"Found {len(boundary_columns)} boundary, {len(target_column_letters)} target columns")

                # Auto-detect province name from the first data row
                update_progress(25, 100, "Detecting province...")
                province_col = boundary_columns.get(2)
                province_name = ''
                if province_col:
                    wb_tmp = openpyxl.load_workbook(file_state['boundary_file'], data_only=True, read_only=True)
                    ws_tmp = None
                    for s in wb_tmp.worksheets:
                        if s.sheet_state == 'visible':
                            ws_tmp = s
                            break
                    if ws_tmp is None:
                        ws_tmp = wb_tmp.active
                    for row in ws_tmp.iter_rows(min_row=header_row + 1, max_row=header_row + 1):
                        for cell in row:
                            if get_column_letter(cell.column) == province_col and cell.value:
                                province_name = str(cell.value).strip()
                                break
                    wb_tmp.close()

                if not province_name:
                    display(HTML(
                        "<p style='color:red'><b>Could not detect province name from boundary file.</b></p>"))
                    btn_transform.disabled = False
                    btn_transform.description = 'TRANSFORM'
                    return

                update_progress(30, 100, f"Province: {province_name}")

                # Format campaign dates
                date_fmt = '%d/%m/%Y'
                campaign_start = w_campaign_start.value.strftime(date_fmt)
                campaign_end = w_campaign_end.value.strftime(date_fmt)
                update_progress(35, 100, "Configuration ready")

                # DB path inside output folder
                db_filename = w_db_name.value.strip() or 'microplan.db'
                db_path_in_output = os.path.join(OUTPUT_DIR, db_filename)

                # Build user inputs
                user_inputs = {
                    'db_name': db_path_in_output,
                    'country_code': w_country_code.value.strip() or 'mz',
                    'province_name': province_name,
                    'province_code': '',
                    'project_name': w_project_name.value.strip() or 'IRS',
                    'boundary_start_row': header_row,
                    'facility_start_row': w_facility_start_row.value,
                    'campaign_start_date': campaign_start,
                    'campaign_end_date': campaign_end,
                }

                config_with_cols = dict(config_state)
                config_with_cols['boundary_columns'] = boundary_columns
                config_with_cols['target_column_letters'] = target_column_letters

                # Reload modules (clears stale state)
                update_progress(40, 100, "Reloading modules...")

                if 'models.db.Base' in sys.modules:
                    sys.modules['models.db.Base'].Base.metadata.clear()

                _modules_to_reload = [
                    'constants.constants',
                    'models.db.Base',
                    'models.db.Boundary',
                    'models.db.Facility',
                    'models.db',
                    'models',
                    'utils.common',
                    'utils.boundary',
                    'utils.facility',
                    'utils',
                    'transform',
                ]
                for mod_name in _modules_to_reload:
                    if mod_name in sys.modules:
                        importlib.reload(sys.modules[mod_name])

                # Restore casing mode after reload
                from utils.common import set_casing_mode
                set_casing_mode(config_state.get('casing_mode', 'none'))

                # Build and apply config AFTER reload
                update_progress(45, 100, "Applying configuration...")
                from constants.constants import TransformConfig
                cfg = TransformConfig.from_notebook(config_with_cols, user_inputs)
                cfg.apply_to_module()

                # Import and run
                from transform import run_transform

                update_progress(50, 100, "Transforming data...")
                result = run_transform(
                    boundary_file=file_state['boundary_file'],
                    facility_file=file_state['facility_file'],
                    progress=update_progress,
                )

                # Show results + download
                db_path = result['db_path']
                html = f"""
                <div style='padding:12px; background:#d4edda; border-left:4px solid #27ae60; margin:10px 0;'>
                    <b style='color:#27ae60; font-size:16px;'>Transformation Complete!</b><br><br>
                    <table style='font-size:13px;'>
                        <tr><td><b>Database:</b></td><td>{db_path}</td></tr>
                        <tr><td><b>Province:</b></td><td>{province_name}</td></tr>
                        <tr><td><b>Campaign:</b></td><td>{campaign_start} &rarr; {campaign_end}</td></tr>
                        <tr><td><b>Boundaries created:</b></td><td>{result['boundaries_count']}</td></tr>
                        <tr><td><b>Facilities created:</b></td><td>{result['facilities_count']}</td></tr>
                    </table>
                </div>"""

                if os.path.exists(db_path):
                    with open(db_path, 'rb') as f:
                        b64 = base64.b64encode(f.read()).decode()
                    html += (
                        f'<a href="data:application/octet-stream;base64,{b64}" '
                        f'download="{os.path.basename(db_path)}" '
                        f'style="display:inline-block; padding:10px 20px; background:#2196F3; color:white; '
                        f'text-decoration:none; border-radius:4px; margin:8px 4px; font-weight:bold;">'
                        f'Download {os.path.basename(db_path)}</a>'
                    )

                progress_bar.bar_style = 'success'
                progress_label.value = (
                    "<span style='margin-left:10px; color:#2e7d32; font-weight:bold;'>"
                    "Transformation complete!</span>"
                )
                display(HTML(html))

            except Exception as ex:
                import traceback
                progress_bar.bar_style = 'danger'
                progress_label.value = (
                    "<span style='margin-left:10px; color:#c62828; font-weight:bold;'>"
                    "Failed</span>"
                )
                display(HTML(
                    f"<div style='padding:10px; background:#ffc7ce; border-radius:5px;'>"
                    f"<b style='color:red'>Transformation Error:</b><br>"
                    f"<pre>{traceback.format_exc()}</pre></div>"
                ))
            finally:
                btn_transform.disabled = False
                btn_transform.description = 'TRANSFORM'

    btn_transform.on_click(on_transform)

    return widgets.VBox([
        widgets.HTML(info_html),
        out_transform_status,
        widgets.HTML("<b>Settings:</b>"),
        w_db_name, w_country_code, w_project_name,
        widgets.HTML("<b>Campaign Dates:</b>"),
        w_campaign_start, w_campaign_end,
        widgets.HTML("<b>Excel Row Settings:</b>"),
        w_boundary_start_row, w_facility_start_row,
        widgets.HTML("<br>"),
        btn_transform,
        progress_box,
        out_transform,
    ])
