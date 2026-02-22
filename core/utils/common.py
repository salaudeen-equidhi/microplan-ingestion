import os
import csv

import shortuuid
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import constants


def generate_short_code() -> str:
    return str(shortuuid.uuid())


def convert_to_date_format(date_to_be_converted, converted_format=constants.DATE_FORMAT) -> str:
    return date_to_be_converted.strftime(converted_format)


def is_excel(filename: os.DirEntry) -> bool:
    extension = os.path.splitext(filename)[1][1:]
    if extension in ["xlsx", "xls"]:
        return True
    return False


def get_visible_sheets(sheets: [Worksheet]):
    visible_sheets = []
    for sheet in sheets:
        if sheet.sheet_state == "visible":
            visible_sheets.append(sheet)
    return visible_sheets


def detect_columns(file_path, header_row=1):
    """Read header row from Excel or CSV -> {column_name: column_letter} mapping."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        mapping = {}
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader, start=1):
                if i == header_row:
                    for col_idx, val in enumerate(row):
                        if val and val.strip():
                            col_letter = get_column_letter(col_idx + 1)
                            mapping[val.strip()] = col_letter
                    break
        return mapping
    # Excel (.xlsx, .xls)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = None
    for s in wb.worksheets:
        if s.sheet_state == 'visible':
            ws = s
            break
    if ws is None:
        ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell.value is not None:
                col_letter = get_column_letter(cell.column)
                mapping[str(cell.value).strip()] = col_letter
    wb.close()
    return mapping


def classify_columns(file_path, header_row=1, sample_rows=20):
    """Classify boundary file columns as text (levels) or numeric (targets).
    Returns (level_cols, target_cols) as two lists of header names, in order."""
    ext = os.path.splitext(file_path)[1].lower()
    headers = []
    data_rows = []

    if ext == '.csv':
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader, start=1):
                if i == header_row:
                    headers = [v.strip() for v in row if v and v.strip()]
                elif i > header_row:
                    data_rows.append(row)
                    if len(data_rows) >= sample_rows:
                        break
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = None
        for s in wb.worksheets:
            if s.sheet_state == 'visible':
                ws = s
                break
        if ws is None:
            ws = wb.active
        for row in ws.iter_rows(min_row=header_row, max_row=header_row + sample_rows):
            row_vals = [cell.value for cell in row]
            if not headers:
                headers = [str(v).strip() if v is not None else '' for v in row_vals]
            else:
                data_rows.append(row_vals)
        wb.close()

    if not headers or not data_rows:
        return headers, []

    level_cols = []
    target_cols = []
    for col_idx, name in enumerate(headers):
        if not name:
            continue
        numeric_count = 0
        non_empty = 0
        for row in data_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                val = row[col_idx]
                val_str = str(val).strip()
                if val_str:
                    non_empty += 1
                    try:
                        float(val)
                        numeric_count += 1
                    except (ValueError, TypeError):
                        pass
        if non_empty > 0 and numeric_count / non_empty > 0.5:
            target_cols.append(name)
        else:
            level_cols.append(name)

    return level_cols, target_cols


CASING_OPTIONS = {
    'none': 'No change (keep original)',
    'title': 'Title Case (Each Word Capitalized)',
    'upper': 'UPPER CASE',
    'lower': 'lower case',
    'sentence': 'Sentence case (first word only)',
}

# Active casing mode — set by config UI, used by cleanup()
_casing_mode = 'none'


def set_casing_mode(mode):
    global _casing_mode
    _casing_mode = mode


def cleanup(name):
    """Normalize name: strip whitespace and apply chosen casing."""
    if name is None:
        return name
    s = str(name).strip()
    if _casing_mode == 'title':
        return s.title()
    elif _casing_mode == 'upper':
        return s.upper()
    elif _casing_mode == 'lower':
        return s.lower()
    elif _casing_mode == 'sentence':
        return s.capitalize()
    return s
