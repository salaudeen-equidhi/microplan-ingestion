import logging
import os

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import constants.constants as constants
import utils.common
import utils.boundary
import utils.facility
from models.db import Base, Boundary, Facility
from utils.common import cleanup


def run_transform(boundary_file, facility_file, checklist_file=None, progress=None):
    """Run the microplan transformation pipeline.

    Args:
        boundary_file: Path to boundary Excel file or directory containing them.
        facility_file: Path to facility Excel file or directory containing them.
        checklist_file: Optional path to checklist_targets.xlsx.
        progress: Optional callback function(message: str) for progress updates.

    Returns:
        dict with db_path, boundaries_count, facilities_count.
    """
    def log(msg):
        if progress:
            progress(msg)

    db_url = constants.DB_CONNECTION_STRING
    log("Connecting to database...")

    engine = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    session = Session()
    Base.metadata.create_all(engine)
    log("Database ready.")

    # ── Checklist targets ──────────────────────────────────────
    checklist_targets_json = {}
    if checklist_file and os.path.exists(checklist_file):
        log("Loading checklist targets...")
        wb = openpyxl.load_workbook(checklist_file)
        sheet = wb.active
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        col_names = header[1:]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            bname, *tgts = row
            if not bname:
                continue
            key = cleanup(bname)
            if key not in checklist_targets_json:
                checklist_targets_json[key] = {"targets": []}
            for cn, tv in zip(col_names, tgts):
                if cn:
                    checklist_targets_json[key]["targets"].append(
                        {"beneficiaryType": cn, "totalNo": tv, "targetNo": tv})

    # ── Max boundary level (for target extraction) ─────────────
    max_level = max(info["level"] for info in constants.BOUNDARIES.values())

    # ── BOUNDARY_1 ─────────────────────────────────────────────
    boundaries = {}
    boundary_1 = Boundary(code=constants.BOUNDARY_1_CODE)
    boundaries['BOUNDARY_1'] = boundary_1

    # ── Resolve boundary input files ───────────────────────────
    if os.path.isfile(boundary_file):
        b_files = [boundary_file]
    else:
        b_files = [f.path for f in os.scandir(boundary_file)
                   if f.is_file() and utils.common.is_excel(f)]

    # ── Create boundaries ──────────────────────────────────────
    for fpath in b_files:
        log(f"Loading boundary file: {os.path.basename(fpath)}...")
        wb = openpyxl.load_workbook(fpath, data_only=True)
        sheets = utils.common.get_visible_sheets(wb.worksheets)
        log(f"Found {len(sheets)} sheet(s). Creating boundaries...")

        b2 = utils.boundary.upsert_boundary_2(
            constants.BOUNDARY_2_NAME, constants.BOUNDARY_2_CODE,
            "BOUNDARY_2", session, str(fpath))
        boundaries["BOUNDARY_2"] = b2

        for si, ws in enumerate(sheets, 1):
            last_row = ws.max_row  # snapshot before loop to avoid openpyxl expanding
            total_rows = last_row - constants.START_BOUNDARIES_ROW
            log(f"Processing sheet {si}/{len(sheets)}: {ws.title} ({total_rows} rows)...")
            row = constants.START_BOUNDARIES_ROW
            processed = 0
            while row < last_row + 1:
                row += 1
                processed += 1
                if processed % 100 == 0:
                    log(f"  Boundaries: {processed}/{total_rows} rows processed...")
                targets = {}

                # Extract targets at the deepest boundary level
                for bk, bi in constants.BOUNDARIES.items():
                    if bi["level"] >= max_level:
                        col = bi.get("column")
                        if col and ws[f"{col}{row}"].value is not None:
                            for tn, tc in constants.TARGET_COLUMNS.items():
                                targets[tn] = ws[f"{tc}{row}"].value

                # Create boundary hierarchy (level 3+)
                for bk, bi in constants.BOUNDARIES.items():
                    if bi["level"] >= 3:
                        col = bi.get("column")
                        if col and ws[f"{col}{row}"].value is not None:
                            bname = cleanup(ws[f"{col}{row}"].value)
                            btype = bi["name"]
                            prev_key = constants.get_boundary_name(bi["level"] - 1)
                            b = utils.boundary.upsert_boundary(
                                bname, bk, boundaries[prev_key].code,
                                session, btype, str(fpath), targets)
                            boundaries[bk] = b

    # ── Resolve facility input files ───────────────────────────
    if os.path.isfile(facility_file):
        f_files = [facility_file]
    else:
        f_files = [f.path for f in os.scandir(facility_file)
                   if f.is_file() and utils.common.is_excel(f)]

    # ── Create facilities ──────────────────────────────────────
    for fpath in f_files:
        log(f"Loading facility file: {os.path.basename(fpath)}...")
        wb = openpyxl.load_workbook(fpath, data_only=True)
        sheets = utils.common.get_visible_sheets(wb.worksheets)
        log(f"Found {len(sheets)} sheet(s). Creating facilities...")

        for ws in sheets:
            total_fac_rows = ws.max_row - constants.FACILITY_START_ROW
            fac_processed = 0
            for row in range(constants.FACILITY_START_ROW + 1, ws.max_row + 1):
                fac_processed += 1
                if fac_processed % 100 == 0:
                    log(f"  Facilities: {fac_processed}/{total_fac_rows} rows processed...")
                cells = ws[f"A{row}:F{row}"]
                vals = [c.value for c in cells[0]]
                fac_name = vals[0]
                mapping = vals[2]
                if not fac_name or not mapping:
                    continue
                utils.facility.create_health_facility(
                    facility_name=fac_name,
                    mapping_boundary=mapping,
                    session=session,
                    facility_type="Health Facility",
                    filename=str(fpath),
                    target=0)

    # ── Results ────────────────────────────────────────────────
    log("Finalizing database...")
    b_count = session.query(Boundary).count()
    f_count = session.query(Facility).count()
    db_path = db_url.replace("sqlite:///", "")
    log(f"Done! {b_count} boundaries, {f_count} facilities created.")

    session.close()
    engine.dispose()

    return {
        'db_path': db_path,
        'boundaries_count': b_count,
        'facilities_count': f_count,
    }


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description="Transform microplan to DIGIT Health Ingestion format",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("-v", "--verbose", action="store_true")
    parser.add_argument("-bp", "--boundary-path",
                        default="./files/input/boundary/data-1/")
    parser.add_argument("-hfp", "--hf-path",
                        default="./files/input/facility/data-1/")
    args = parser.parse_args()

    if args.verbose:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s : %(levelname)s : %(message)s",
            datefmt="%Y-%m-%d %I:%M:%S%p")

    cl = "checklist_targets.xlsx" if os.path.exists("checklist_targets.xlsx") else None
    result = run_transform(args.boundary_path, args.hf_path, cl)
    print(f"Done! DB: {result['db_path']}, "
          f"Boundaries: {result['boundaries_count']}, "
          f"Facilities: {result['facilities_count']}")
