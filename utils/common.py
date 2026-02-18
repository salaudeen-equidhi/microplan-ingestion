import os
import csv

import shortuuid
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import constants


def generate_short_code() -> str:
    return str(shortuuid.uuid())


def convert_to_date_format(date_to_be_converted, converted_format=constants.DATE_FORMAT) -> str:
    return date_to_be_converted.strftime(converted_format)


def is_excel(filename: os.DirEntry) -> bool:
    extension = os.path.splitext(filename)[1][1:]
    if extension in ["xlsx", "xls"]:
        return True
    return False


def get_visible_sheets(sheets: [Worksheet]):
    visible_sheets = []
    for sheet in sheets:
        if sheet.sheet_state == "visible":
            visible_sheets.append(sheet)
    return visible_sheets


def detect_columns(file_path, header_row=1):
    """Read header row from Excel or CSV -> {column_name: column_letter} mapping."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        mapping = {}
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader, start=1):
                if i == header_row:
                    for col_idx, val in enumerate(row):
                        if val and val.strip():
                            col_letter = get_column_letter(col_idx + 1)
                            mapping[val.strip()] = col_letter
                    break
        return mapping
    # Excel (.xlsx, .xls)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = None
    for s in wb.worksheets:
        if s.sheet_state == 'visible':
            ws = s
            break
    if ws is None:
        ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell.value is not None:
                col_letter = get_column_letter(cell.column)
                mapping[str(cell.value).strip()] = col_letter
    wb.close()
    return mapping


def cleanup(name):
    return name
